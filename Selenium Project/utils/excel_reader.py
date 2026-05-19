'''import os
from openpyxl import load_workbook


class ExcelReader:

    @staticmethod
    def read_excel(file_name, sheet_name):
        data = []

        base_dir = os.path.dirname(os.path.dirname(__file__))
        file_path = str(os.path.join(base_dir, "data", file_name))

        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]

        headers = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))
            data.append(row_data)

        return data'''
import openpyxl


class ExcelReader:

    @staticmethod
    def read_excel(file_path, sheet_name):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        data = []
        headers = []

        # READ HEADERS & CLEAN HIDDEN SPACES
        for cell in sheet[1]:
            if cell.value:
                # .strip() removes any accidental spaces at the start/end of the header text
                headers.append(str(cell.value).strip())
            else:
                headers.append(None)

        # READ ROWS (Changed to sheet.max_row so it can read multiple test cases)
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):

            # Skip entirely empty rows at the bottom of the sheet
            if all(cell is None for cell in row):
                continue

            row_data = {}
            for key, value in zip(headers, row):
                if key:  # Ensure we don't map None keys
                    row_data[key] = value

            data.append(row_data)

        workbook.close()
        return data